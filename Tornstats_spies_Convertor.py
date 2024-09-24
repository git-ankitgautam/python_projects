from openpyxl import Workbook, load_workbook
workbook = load_workbook('Torn_Stats_-_Personal_Spies-2 (1).xlsx')
worksheet = workbook['Sheet1']

def add_to_file(nameID,lvl,Speed,Strength,Dexterity,Defense,statTotal):
    with open("testSpies2.txt",'a') as f:
            f.write("Name: " + nameID +"\n")
            f.write("Level: " + str(lvl) +"\n\n" + "You managed to get the following results:\n")
            f.write("Speed: " + str(Speed) +"\n")
            f.write("Strength: " + str(Strength) + "\n")
            f.write("Defense: " + str(Defense)+ "\n")
            f.write("Dexterity: " + str(Dexterity)+ "\n")
            f.write("Total: " + str(statTotal)+ "\n\n")

max = worksheet.max_row
for row in worksheet.iter_rows(min_row=2, max_row=16):
    nameID = row[1].value
    lvl = int(row[2].value)
    Speed = int(row[6].value)
    Strength = int(row[4].value)
    Dexterity = int(row[7].value)
    Defense = int(row[5].value)
    statTotal = int(row[8].value)
    add_to_file(nameID,lvl,Speed,Strength,Dexterity,Defense,statTotal)


